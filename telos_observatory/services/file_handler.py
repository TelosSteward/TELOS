"""
File Handler Service for TELOS Observatory.

Handles file uploads, extraction, and processing for chat context.
Supports: PDF, images, text, code, documents, spreadsheets.
"""

import io
import base64
import mimetypes
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import streamlit as st


class FileHandler:
    """Handle file uploads and extraction for TELOS chat."""

    # Supported file types
    SUPPORTED_TYPES = {
        # Documents
        'pdf': ['application/pdf'],
        'txt': ['text/plain'],
        'md': ['text/markdown', 'text/x-markdown'],
        'doc': ['application/msword'],
        'docx': ['application/vnd.openxmlformats-officedocument.wordprocessingml.document'],

        # Images
        'png': ['image/png'],
        'jpg': ['image/jpeg'],
        'jpeg': ['image/jpeg'],
        'gif': ['image/gif'],
        'webp': ['image/webp'],
        'svg': ['image/svg+xml'],

        # Code
        'py': ['text/x-python'],
        'js': ['text/javascript', 'application/javascript'],
        'jsx': ['text/jsx'],
        'ts': ['text/typescript'],
        'tsx': ['text/tsx'],
        'html': ['text/html'],
        'css': ['text/css'],
        'json': ['application/json'],
        'xml': ['application/xml', 'text/xml'],
        'yaml': ['application/x-yaml', 'text/yaml'],
        'yml': ['application/x-yaml', 'text/yaml'],

        # Spreadsheets
        'csv': ['text/csv'],
        'xlsx': ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'],
        'xls': ['application/vnd.ms-excel'],
    }

    # Maximum file size (10 MB)
    MAX_FILE_SIZE = 10 * 1024 * 1024

    def __init__(self):
        """Initialize file handler."""
        self.uploaded_files = []

    def process_uploaded_file(self, uploaded_file) -> Dict[str, Any]:
        """
        Process a single uploaded file.

        Args:
            uploaded_file: Streamlit UploadedFile object

        Returns:
            Dictionary with file metadata and extracted content
        """
        try:
            # Get file info
            file_name = uploaded_file.name
            file_size = uploaded_file.size
            file_type = uploaded_file.type

            # Check file size
            if file_size > self.MAX_FILE_SIZE:
                return {
                    'name': file_name,
                    'size': file_size,
                    'type': file_type,
                    'error': f'File too large ({file_size / 1024 / 1024:.1f}MB). Max size is 10MB.',
                    'success': False
                }

            # Read file bytes
            file_bytes = uploaded_file.read()

            # Extract text content based on file type
            content, error = self._extract_content(file_bytes, file_name, file_type)

            if error:
                return {
                    'name': file_name,
                    'size': file_size,
                    'type': file_type,
                    'error': error,
                    'success': False
                }

            # Prepare file info
            file_info = {
                'name': file_name,
                'size': file_size,
                'type': file_type,
                'content': content,
                'bytes': file_bytes,
                'base64': base64.b64encode(file_bytes).decode('utf-8'),
                'success': True,
                'is_image': self._is_image(file_type),
                'is_text': self._is_text(file_type),
            }

            return file_info

        except Exception as e:
            return {
                'name': uploaded_file.name if hasattr(uploaded_file, 'name') else 'unknown',
                'error': 'Error processing file. Please try a different file.',
                'success': False
            }

    def _extract_content(self, file_bytes: bytes, file_name: str, file_type: str) -> Tuple[Optional[str], Optional[str]]:
        """
        Extract text content from file bytes.

        Args:
            file_bytes: File content as bytes
            file_name: Original file name
            file_type: MIME type

        Returns:
            Tuple of (content, error)
        """
        try:
            # Get file extension
            ext = Path(file_name).suffix.lower().lstrip('.')

            # PDF extraction
            if ext == 'pdf' or file_type == 'application/pdf':
                return self._extract_pdf(file_bytes)

            # Image - no text extraction, return metadata
            if self._is_image(file_type):
                return f"[Image: {file_name}]", None

            # Text-based files
            if self._is_text(file_type) or ext in ['txt', 'md', 'py', 'js', 'html', 'css', 'json', 'xml', 'yaml', 'yml', 'csv']:
                try:
                    content = file_bytes.decode('utf-8')
                    return content, None
                except UnicodeDecodeError:
                    try:
                        content = file_bytes.decode('latin-1')
                        return content, None
                    except:
                        return None, "Could not decode text file"

            # DOCX extraction
            if ext == 'docx':
                return self._extract_docx(file_bytes)

            # XLSX extraction
            if ext == 'xlsx':
                return self._extract_xlsx(file_bytes)

            # Unsupported type
            return None, f"Unsupported file type: {file_type}"

        except Exception as e:
            return None, "Error extracting file content. Please try a different file."

    def _extract_pdf(self, file_bytes: bytes) -> Tuple[Optional[str], Optional[str]]:
        """Extract text from PDF file."""
        try:
            import PyPDF2
            pdf_file = io.BytesIO(file_bytes)
            reader = PyPDF2.PdfReader(pdf_file)

            text_parts = []
            for page in reader.pages:
                text_parts.append(page.extract_text())

            content = "\n\n".join(text_parts)
            return content, None

        except ImportError:
            return None, "PDF support requires PyPDF2. Install with: pip install PyPDF2"
        except Exception as e:
            return None, "Error extracting PDF content. Please try a different file."

    def _extract_docx(self, file_bytes: bytes) -> Tuple[Optional[str], Optional[str]]:
        """Extract text from DOCX file."""
        try:
            import docx
            doc_file = io.BytesIO(file_bytes)
            doc = docx.Document(doc_file)

            text_parts = []
            for paragraph in doc.paragraphs:
                text_parts.append(paragraph.text)

            content = "\n\n".join(text_parts)
            return content, None

        except ImportError:
            return None, "DOCX support requires python-docx. Install with: pip install python-docx"
        except Exception as e:
            return None, "Error extracting DOCX content. Please try a different file."

    def _extract_xlsx(self, file_bytes: bytes) -> Tuple[Optional[str], Optional[str]]:
        """Extract text from XLSX file."""
        try:
            import openpyxl
            xlsx_file = io.BytesIO(file_bytes)
            workbook = openpyxl.load_workbook(xlsx_file)

            text_parts = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"Sheet: {sheet_name}\n")

                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    text_parts.append(row_text)

                text_parts.append("\n")

            content = "\n".join(text_parts)
            return content, None

        except ImportError:
            return None, "XLSX support requires openpyxl. Install with: pip install openpyxl"
        except Exception as e:
            return None, "Error extracting XLSX content. Please try a different file."

    def _is_image(self, file_type: str) -> bool:
        """Check if file is an image."""
        return file_type.startswith('image/')

    def _is_text(self, file_type: str) -> bool:
        """Check if file is text-based."""
        text_types = [
            'text/', 'application/json', 'application/xml',
            'application/javascript', 'application/x-yaml'
        ]
        return any(file_type.startswith(t) for t in text_types)

    def format_file_context(self, files: List[Dict[str, Any]]) -> str:
        """
        Format uploaded files for LLM context.

        Args:
            files: List of processed file dictionaries

        Returns:
            Formatted string for LLM prompt
        """
        if not files:
            return ""

        context_parts = ["\n\n### Attached Files:\n"]

        for file_info in files:
            if not file_info.get('success'):
                continue

            file_name = file_info['name']
            file_type = file_info['type']
            content = file_info.get('content', '')

            if file_info.get('is_image'):
                context_parts.append(f"\n**File:** {file_name} (Image)\n")
                context_parts.append(f"*[Image file attached - visual content available]*\n")
            elif content:
                context_parts.append(f"\n**File:** {file_name}\n")
                context_parts.append(f"```\n{content}\n```\n")

        return "".join(context_parts)

    def render_file_upload(self, key_suffix: str = "main") -> List[Dict[str, Any]]:
        """
        Render file upload widget and return processed files.

        Args:
            key_suffix: Unique suffix for the upload widget key

        Returns:
            List of processed file dictionaries
        """
        uploaded_files = st.file_uploader(
            "Attach files (optional)",
            type=None,  # Allow all types
            accept_multiple_files=True,
            key=f"file_upload_{key_suffix}",
            label_visibility="collapsed"
        )

        if not uploaded_files:
            return []

        processed_files = []
        for uploaded_file in uploaded_files:
            file_info = self.process_uploaded_file(uploaded_file)
            processed_files.append(file_info)

        return processed_files

    def render_uploaded_files(self, files: List[Dict[str, Any]]):
        """
        Render previews of uploaded files.

        Args:
            files: List of processed file dictionaries
        """
        if not files:
            return

        st.markdown("**Attached Files:**")

        for file_info in files:
            if not file_info.get('success'):
                st.error(f"❌ {file_info.get('name', 'Unknown')}: {file_info.get('error', 'Unknown error')}")
                continue

            file_name = file_info['name']
            file_size = file_info['size']
            file_type = file_info['type']

            # Format size
            if file_size < 1024:
                size_str = f"{file_size}B"
            elif file_size < 1024 * 1024:
                size_str = f"{file_size / 1024:.1f}KB"
            else:
                size_str = f"{file_size / 1024 / 1024:.1f}MB"

            # File icon based on type
            if file_info.get('is_image'):
                icon = "🖼️"
            elif file_type == 'application/pdf':
                icon = "📄"
            elif 'spreadsheet' in file_type or file_name.endswith(('.csv', '.xlsx', '.xls')):
                icon = "📊"
            elif 'word' in file_type or file_name.endswith(('.doc', '.docx')):
                icon = "📝"
            else:
                icon = "📎"

            # Render file info
            col1, col2 = st.columns([4, 1])
            with col1:
                st.markdown(f"{icon} **{file_name}**")
            with col2:
                st.markdown(f"*{size_str}*")

            # Show image preview if it's an image
            if file_info.get('is_image'):
                st.image(file_info['bytes'], caption=file_name, use_container_width=True)

            # Show text preview for small text files
            elif file_info.get('is_text') and file_info.get('content'):
                content = file_info['content']
                if len(content) < 500:
                    with st.expander("Preview"):
                        st.code(content, language='text')


# Global instance
_file_handler = None


def get_file_handler() -> FileHandler:
    """Get or create file handler instance."""
    global _file_handler
    if _file_handler is None:
        _file_handler = FileHandler()
    return _file_handler
